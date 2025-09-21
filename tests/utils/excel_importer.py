import os
from openpyxl import load_workbook
from django.conf import settings
from tests.models import Test, Question

def import_test_from_excel(file_path, test_name=None):
    """
    Импортирует тест из Excel файла
    Формат файла:
    - Столбец A: номер вопроса
    - Столбец B: текст вопроса
    - Столбец C: правильные ответы (могут быть несколько через запятую)
    - Столбец D: ссылка на документ
    - Столбцы E и далее: варианты ответов
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Файл {file_path} не найден")
    
    # Создаем или получаем тест
    if test_name is None:
        test_name = os.path.basename(file_path).split('.')[0]
    
    test, created = Test.objects.get_or_create(name=test_name)
    if created:
        test.description = f"Тест импортирован из файла {os.path.basename(file_path)}"
        test.save()
    
    # Очищаем существующие вопросы
    test.questions.all().delete()
    
    # Загружаем Excel файл
    wb = load_workbook(filename=file_path)
    sheet = wb.active
    
    # Пропускаем заголовок если есть
    start_row = 2 if sheet['A1'].value and 'опрос' in str(sheet['A1'].value) else 1
    
    questions_created = 0
    for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not row[0] or not isinstance(row[0], (int, float)):
            continue  # Пропускаем пустые строки или строки без номера вопроса
        
        question_number = int(row[0])
        question_text = row[1] if row[1] else ""
        
        # Обрабатываем правильные ответы (могут быть несколько через запятую)
        correct_answer = str(row[2]) if row[2] else "1"
        
        document_reference = row[3] if row[3] else ""
        
        # Собираем варианты ответов
        answer_options = {}
        option_num = 1
        for i in range(4, len(row)):
            if row[i] and str(row[i]).strip():
                answer_options[option_num] = str(row[i])
                option_num += 1
        
        # Проверяем, что есть варианты ответов
        if not answer_options:
            raise ValueError(f"В строке {row_num} нет вариантов ответов")
        
        # Создаем вопрос
        Question.objects.create(
            test=test,
            question_number=question_number,
            question_text=question_text,
            correct_answer=correct_answer,
            document_reference=document_reference,
            answer_options=answer_options
        )
        questions_created += 1
    
    if questions_created == 0:
        raise ValueError("Не удалось импортировать ни одного вопроса из файла")
    
    return test